import pandas as pd
import os
import csv
import io
import logging
from typing import Dict, Tuple, List, Optional
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.utils import get_column_letter

class FileHandler:
    @staticmethod
    def read_file(file_path: str) -> Dict[str, pd.DataFrame]:
        """Read file and return dictionary of DataFrames by sheet name - from original app.py"""
        try:
            logging.debug(f"Reading file: {file_path}")
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                xl = pd.ExcelFile(file_path)
                logging.debug(f"Excel file detected, sheets: {xl.sheet_names}")
                sheets = {sheet_name: pd.read_excel(file_path, sheet_name=sheet_name, header=None) 
                         for sheet_name in xl.sheet_names}
                return sheets
            elif file_path.endswith(('.txt', '.csv', '.dat')):
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    content = f.read()
                if not content.strip():
                    logging.error("File is empty")
                    raise ValueError("File is empty.")
                try:
                    dialect = csv.Sniffer().sniff(content[:1024])
                    sep = dialect.delimiter
                    logging.debug(f"CSV file detected, delimiter: {sep}")
                except:
                    sep = FileHandler.detect_delimiter(file_path)
                    logging.debug(f"Delimiter detection failed, using fallback: {sep}")
                df = pd.read_csv(file_path, header=None, sep=sep, encoding='utf-8', quotechar='"', engine='python')
                df.columns = [str(col) for col in df.columns]
                logging.debug(f"CSV file read, shape: {df.shape}")
                return {'Sheet1': df}
            else:
                logging.error("Unsupported file type")
                raise ValueError("Unsupported file type.")
        except Exception as e:
            logging.error(f"Error reading file {file_path}: {str(e)}")
            raise ValueError(f"Error reading file: {str(e)}")

    @staticmethod
    def detect_delimiter(file_path: str) -> str:
        """Intelligent delimiter detection for CSV files - from original app.py"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read(1024)
            if not content.strip():
                logging.warning("File content is empty, using default delimiter: ','")
                return ','
            delimiters = [',', ';', '|', '/', '\t', ':', '-']
            best_delimiter, max_columns, best_consistency = None, 0, 0
            for delim in delimiters:
                try:
                    sample_df = pd.read_csv(io.StringIO(content), sep=delim, header=None, nrows=5, quotechar='"', engine='python')
                    column_count = sample_df.shape[1]
                    row_lengths = [len(row.dropna()) for _, row in sample_df.iterrows()]
                    consistency = sum(1 for length in row_lengths if length == column_count) / len(row_lengths)
                    if column_count > 1 and column_count > max_columns and consistency > best_consistency:
                        max_columns = column_count
                        best_consistency = consistency
                        best_delimiter = delim
                except Exception:
                    continue
            delimiter = best_delimiter or ','
            logging.debug(f"Detected delimiter: {delimiter}")
            return delimiter
        except Exception as e:
            logging.error(f"Error detecting delimiter for {file_path}: {str(e)}")
            return ','

    @staticmethod
    def find_header_row(df: pd.DataFrame, max_rows: int = 10) -> int:
        """Intelligent header row detection - from original app.py"""
        try:
            for i in range(min(len(df), max_rows)):
                row = df.iloc[i].dropna()
                if not row.empty and all(isinstance(x, str) for x in row if pd.notna(x)):
                    logging.debug(f"Header row detected at index {i}")
                    return i
            logging.warning(f"No header row detected within the first {max_rows} rows")
            return 0 if not df.empty and len(df.columns) > 0 else -1
        except Exception as e:
            logging.error(f"Error finding header row: {str(e)}")
            return -1

    @staticmethod
    def save_corrected_file(df: pd.DataFrame, original_filename: str, upload_folder: str, 
                           sheet_name: Optional[str] = None, phase: str = "corrected") -> str:
        """Save corrected DataFrame with proper file naming"""
        try:
            base_name, ext = os.path.splitext(original_filename)
            corrected_filename = f"{base_name}_{phase}{ext}"
            corrected_file_path = os.path.join(upload_folder, corrected_filename)
            
            if ext.lower() == '.xlsx':
                df.to_excel(corrected_file_path, index=False, sheet_name=sheet_name or 'Sheet1')
                logging.info(f"Saved Excel file: {corrected_file_path}")
            else:
                df.to_csv(corrected_file_path, index=False)
                logging.info(f"Saved CSV file: {corrected_file_path}")
            
            return corrected_file_path
        except Exception as e:
            logging.error(f"Error saving corrected file: {str(e)}")
            raise

    @staticmethod
    def create_excel_with_formatting(df: pd.DataFrame, file_path: str, sheet_name: str = 'Sheet1'):
        """Create Excel file with proper formatting and error handling"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Write headers
            for col_idx, header in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=str(header))
            
            # Write data with error handling for illegal characters
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    try:
                        # Handle illegal characters in Excel
                        if isinstance(value, str):
                            # Remove illegal characters for XML
                            value = ''.join(char for char in value if ord(char) >= 32 or char in ['\t', '\n', '\r'])
                        ws.cell(row=row_idx, column=col_idx, value=value)
                    except IllegalCharacterError:
                        # Replace with cleaned string
                        ws.cell(row=row_idx, column=col_idx, value=str(value).encode('ascii', 'ignore').decode('ascii'))
            
            wb.save(file_path)
            logging.info(f"Excel file saved successfully: {file_path}")
        except Exception as e:
            logging.error(f"Error creating Excel file: {str(e)}")
            raise

    @staticmethod
    def validate_file_size(file_path: str, max_size_mb: int = 100) -> bool:
        """Validate file size"""
        try:
            size_mb = os.path.getsize(file_path) / (1024 * 1024)
            if size_mb > max_size_mb:
                logging.warning(f"File size {size_mb:.2f}MB exceeds limit {max_size_mb}MB")
                return False
            return True
        except Exception as e:
            logging.error(f"Error validating file size: {str(e)}")
            return False

    @staticmethod
    def validate_file_extension(filename: str, allowed_extensions: List[str] = None) -> bool:
        """Validate file extension"""
        if allowed_extensions is None:
            allowed_extensions = ['.xlsx', '.xls', '.csv', '.txt', '.dat']
        
        file_ext = os.path.splitext(filename)[1].lower()
        return file_ext in allowed_extensions

    @staticmethod
    def get_file_info(file_path: str) -> Dict:
        """Get comprehensive file information"""
        try:
            stat = os.stat(file_path)
            return {
                'path': file_path,
                'filename': os.path.basename(file_path),
                'size_bytes': stat.st_size,
                'size_mb': round(stat.st_size / (1024 * 1024), 2),
                'extension': os.path.splitext(file_path)[1].lower(),
                'modified_time': stat.st_mtime,
                'exists': True
            }
        except Exception as e:
            logging.error(f"Error getting file info: {str(e)}")
            return {'exists': False, 'error': str(e)}

    @staticmethod
    def cleanup_temp_files(directory: str, max_age_hours: int = 24):
        """Clean up temporary files older than specified hours"""
        try:
            import time
            current_time = time.time()
            max_age_seconds = max_age_hours * 3600
            
            for filename in os.listdir(directory):
                file_path = os.path.join(directory, filename)
                if os.path.isfile(file_path):
                    file_age = current_time - os.path.getmtime(file_path)
                    if file_age > max_age_seconds:
                        os.remove(file_path)
                        logging.info(f"Cleaned up temp file: {filename}")
        except Exception as e:
            logging.error(f"Error during temp file cleanup: {str(e)}")

    @staticmethod
    def safe_filename(filename: str) -> str:
        """Create a safe filename by removing/replacing problematic characters"""
        import re
        # Remove or replace problematic characters
        safe_name = re.sub(r'[<>:"/\\|?*]', '_', filename)
        # Remove multiple underscores
        safe_name = re.sub(r'_+', '_', safe_name)
        # Ensure it's not too long
        if len(safe_name) > 200:
            name, ext = os.path.splitext(safe_name)
            safe_name = name[:200-len(ext)] + ext
        return safe_name

    @staticmethod
    def backup_file(source_path: str, backup_dir: str) -> str:
        """Create a backup copy of a file"""
        try:
            import shutil
            from datetime import datetime
            
            os.makedirs(backup_dir, exist_ok=True)
            filename = os.path.basename(source_path)
            name, ext = os.path.splitext(filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"{name}_backup_{timestamp}{ext}"
            backup_path = os.path.join(backup_dir, backup_filename)
            
            shutil.copy2(source_path, backup_path)
            logging.info(f"File backed up to: {backup_path}")
            return backup_path
        except Exception as e:
            logging.error(f"Error creating backup: {str(e)}")
            raise
